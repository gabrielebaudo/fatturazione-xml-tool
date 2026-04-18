"""
counter_update.py

Increments the numinvio counter in the '2026 XML'!K2 cell of the xlsm workbook
after a successful XML export.

Uses openpyxl with keep_vba=True to preserve macros and VBA project data
while writing back the updated counter value.
"""

from __future__ import annotations

import openpyxl

_DEFAULT_YEAR = 2026
CELL_REF = "K2"


def _counter_sheet(year: int | None) -> str:
    return f"{year if year else _DEFAULT_YEAR} XML"


class ConcurrentModificationError(Exception):
    """Raised when the workbook was modified between reading and writing."""
    pass


def increment_numinvio(xlsm_path: str, current_value: int, year: int | None = None) -> int:
    """
    Increment the numinvio counter in '{year} XML'!K2 from current_value to current_value+1.

    Writes to the xlsm using openpyxl with keep_vba=True to preserve macros.

    Args:
        xlsm_path: path to the .xlsm workbook
        current_value: the value currently in K2 (as read by get_numinvio).
                       Used as a safety check: if K2 != current_value, raises
                       ConcurrentModificationError (concurrent edit detected).
        year: invoice year used to locate the counter sheet (e.g. 2026 → '2026 XML').
              Defaults to 2026 if not provided.

    Returns:
        The new value (current_value + 1)

    Raises:
        ConcurrentModificationError: if K2 value changed since it was read
        FileNotFoundError: if xlsm_path doesn't exist
        PermissionError: if the file is locked (e.g., open in Excel)
    """
    sheet_name = _counter_sheet(year)
    try:
        # Load with data_only=True first to read the cached (evaluated) value
        # for the concurrent-modification safety check.  K2 may contain a formula
        # whose raw text differs from the numeric result we were given.
        wb_check = openpyxl.load_workbook(xlsm_path, data_only=True, keep_vba=True)
    except FileNotFoundError:
        raise FileNotFoundError(f"Workbook not found: {xlsm_path!r}")

    actual = wb_check[sheet_name][CELL_REF].value

    # Safely compare as integers, handling int/float stored values.
    try:
        actual_int = int(actual)
    except (TypeError, ValueError):
        actual_int = actual  # will fail the equality check below

    if actual_int != int(current_value):
        raise ConcurrentModificationError(
            f"numinvio changed: expected {current_value}, found {actual}"
        )

    # Now load without data_only so we can write back without stripping formula metadata.
    wb = openpyxl.load_workbook(xlsm_path, keep_vba=True)
    ws = wb[sheet_name]

    new_value = int(current_value) + 1
    ws[CELL_REF].value = new_value

    try:
        wb.save(xlsm_path)
    except PermissionError:
        raise PermissionError(
            "Cannot write to workbook: please close the file in Excel first."
        )

    return new_value
