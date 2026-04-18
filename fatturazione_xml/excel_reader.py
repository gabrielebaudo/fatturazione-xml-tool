"""
excel_reader.py

Reads cached cell values from a .xlsm workbook using openpyxl.

Excel for Mac evaluates formulas correctly but cannot export XML directly.
This module reads the already-computed values that Excel cached when the user
last saved the file, using openpyxl with data_only=True.
"""

from __future__ import annotations

import sys
from typing import Any

import openpyxl

from .xlsm_parser import XmlBinding

# XPath fragments that identify critical required fields.
_REQUIRED_XPATH_FRAGMENTS = (
    "ProgressivoInvio",
    "TipoDocumento",
    "Numero",
    "Denominazione",
    "IdCodice",
)

# XPath suffixes that are legitimately optional (skip silently).
_OPTIONAL_XPATH_SUFFIXES = ("/@versione", "/@versioneSchema")

# XPath fragments for optional string fields (skip silently).
_OPTIONAL_STRING_XPATH_FRAGMENTS = ("Causale", "DatiOrdineAcquisto")


def read_cell_values(
    xlsm_path: str,
    sheet_name: str,
    bindings: list[XmlBinding],
) -> list[tuple[XmlBinding, Any]]:
    """
    Read cached values for all bound cells in the given sheet.

    Returns list of (binding, value) pairs. Value is the Python value
    from openpyxl (str, int, float, datetime.datetime, None, etc.).

    Raises:
        ValueError: if a cell has None value and its xpath identifies a
                    critical required field.
        FileNotFoundError: if xlsm_path doesn't exist.
    """
    try:
        wb = openpyxl.load_workbook(xlsm_path, data_only=True, keep_vba=True)
    except FileNotFoundError:
        raise FileNotFoundError(f"Workbook not found: {xlsm_path!r}")

    ws = wb[sheet_name]

    results: list[tuple[XmlBinding, Any]] = []

    for binding in bindings:
        value = ws[binding.cell_ref].value

        if value is None:
            xpath = binding.xpath

            # Silently skip optional attribute xpaths.
            if any(xpath.endswith(suffix) for suffix in _OPTIONAL_XPATH_SUFFIXES):
                results.append((binding, value))
                continue

            # Silently skip optional fields regardless of declared data type.
            if any(
                fragment in xpath for fragment in _OPTIONAL_STRING_XPATH_FRAGMENTS
            ):
                results.append((binding, value))
                continue

            # Check if this is a critical required field.
            if any(fragment in xpath for fragment in _REQUIRED_XPATH_FRAGMENTS):
                raise ValueError(
                    f"Cell {binding.cell_ref} ({xpath}) has no cached value. "
                    "Please save the workbook in Excel before exporting."
                )

            # All other None values: log a warning but continue.
            print(
                f"WARNING: Cell {binding.cell_ref} ({xpath}) has no cached value.",
                file=sys.stderr,
            )

        results.append((binding, value))

    return results


def get_numinvio(xlsm_path: str, year: int | None = None) -> int:
    """
    Read the current numinvio counter from '{year} XML'!K2.

    Args:
        xlsm_path: path to the .xlsm workbook
        year: invoice year used to find the counter sheet (e.g. 2026 → '2026 XML').
              Defaults to 2026 if not provided.

    Returns the integer value (the current count, before incrementing).

    Raises:
        FileNotFoundError: if xlsm_path doesn't exist.
        ValueError: if K2 is None or not a numeric value.
    """
    try:
        wb = openpyxl.load_workbook(xlsm_path, data_only=True, keep_vba=True)
    except FileNotFoundError:
        raise FileNotFoundError(f"Workbook not found: {xlsm_path!r}")

    counter_sheet = f"{year} XML" if year else "2026 XML"
    ws = wb[counter_sheet]
    value = ws["K2"].value

    if value is None or not isinstance(value, (int, float)):
        raise ValueError(
            f"numinvio not found in '{counter_sheet}'!K2 — workbook may not be initialized"
        )

    return int(value)
